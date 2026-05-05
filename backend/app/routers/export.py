from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import extract
from datetime import datetime
import pandas as pd
import io
from openpyxl.styles import Font, Alignment
from .. import models
from ..database import get_db
from ..config import get_settings

router = APIRouter(prefix="/export", tags=["export"])

@router.get("/monthly")
def export_monthly_sales(month: int, year: int, db: Session = Depends(get_db)):
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="Invalid month")
    
    sales = db.query(models.Sale).filter(
        extract('month', models.Sale.date) == month,
        extract('year', models.Sale.date) == year
    ).all()
    
    if not sales:
        raise HTTPException(status_code=404, detail="No sales found for the specified month")
    
    data = []
    for sale in sales:
        for item in sale.items:
            item_gst = (item.amount * item.gst_percentage) / 100
            item_cgst = item_gst / 2
            item_sgst = item_gst / 2
            
            data.append({
                "Invoice No": sale.invoice_number,
                "Date": sale.date.strftime("%Y-%m-%d"),
                "Customer Name": sale.customer.name,
                "GSTIN": sale.customer.gstin or "",
                "Product Name": item.description,
                "HSN": item.hsn_code or "",
                "Quantity": item.quantity,
                "Rate": item.rate,
                "Amount": item.amount,
                "GST %": item.gst_percentage,
                "CGST": round(item_cgst, 2),
                "SGST": round(item_sgst, 2),
                "Total": round(item.amount + item_cgst + item_sgst, 2)
            })
    
    df = pd.DataFrame(data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sales', startrow=3)
        
        worksheet = writer.sheets['Sales']
        settings = get_settings()
        
        max_col = len(df.columns)
        
        # Merge header rows
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        
        # Row 1: Shop name
        cell1 = worksheet.cell(row=1, column=1)
        cell1.value = settings.business_name
        cell1.font = Font(size=14, bold=True)
        cell1.alignment = Alignment(horizontal='center')
        
        # Row 2: Month-Year
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 
                       'July', 'August', 'September', 'October', 'November', 'December']
        cell2 = worksheet.cell(row=2, column=1)
        cell2.value = f"{month_names[month]} {year} Sales"
        cell2.font = Font(size=12, bold=True)
        cell2.alignment = Alignment(horizontal='center')
        
        # Add totals row
        start_row = 5  # data starts from row 5
        data_rows = len(df)
        last_data_row = start_row + data_rows - 1
        total_row = last_data_row + 1
        
        # Column mapping based on dataframe order
        # Invoice No(1), Date(2), Customer(3), GSTIN(4), Product(5), HSN(6), Qty(7), Rate(8), Amount(9), GST%(10), CGST(11), SGST(12), Total(13)
        
        # Add "TOTAL" label in Product Name column (column 5)
        worksheet.cell(row=total_row, column=5).value = "TOTAL"
        worksheet.cell(row=total_row, column=5).font = Font(bold=True)
        
        # Merge TOTAL label across Product Name, HSN, Qty, Rate columns (5-8)
        worksheet.merge_cells(start_row=total_row, start_column=5, end_row=total_row, end_column=8)
        
        # Add formulas for Amount, CGST, SGST, Total
        worksheet.cell(row=total_row, column=9).value = f"=SUM(I{start_row}:I{last_data_row})"
        worksheet.cell(row=total_row, column=9).font = Font(bold=True)
        
        worksheet.cell(row=total_row, column=11).value = f"=SUM(K{start_row}:K{last_data_row})"
        worksheet.cell(row=total_row, column=11).font = Font(bold=True)
        
        worksheet.cell(row=total_row, column=12).value = f"=SUM(L{start_row}:L{last_data_row})"
        worksheet.cell(row=total_row, column=12).font = Font(bold=True)
        
        worksheet.cell(row=total_row, column=13).value = f"=SUM(M{start_row}:M{last_data_row})"
        worksheet.cell(row=total_row, column=13).font = Font(bold=True)
        
        # Auto column width (skipped due to merged cells)
        # for col in worksheet.columns:
        #     max_length = 0
        #     for cell in col:
        #         try:
        #             if cell.value and hasattr(cell, 'column_letter'):
        #                 max_length = max(max_length, len(str(cell.value)))
        #         except:
        #             pass
        #     if max_length > 0:
        #         worksheet.column_dimensions[col[0].column_letter].width = max_length + 2
    
    output.seek(0)
    
    filename = f"sales_{year}_{month:02d}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/all")
def export_all_sales(db: Session = Depends(get_db)):
    sales = db.query(models.Sale).order_by(models.Sale.date.desc()).all()
    
    if not sales:
        raise HTTPException(status_code=404, detail="No sales found")
    
    data = []
    for sale in sales:
        for item in sale.items:
            item_gst = (item.amount * item.gst_percentage) / 100
            item_cgst = item_gst / 2
            item_sgst = item_gst / 2
            
            data.append({
                "Invoice No": sale.invoice_number,
                "Date": sale.date.strftime("%Y-%m-%d"),
                "Customer Name": sale.customer.name,
                "GSTIN": sale.customer.gstin or "",
                "Product Name": item.description,
                "HSN": item.hsn_code or "",
                "Quantity": item.quantity,
                "Rate": item.rate,
                "Amount": item.amount,
                "GST %": item.gst_percentage,
                "CGST": round(item_cgst, 2),
                "SGST": round(item_sgst, 2),
                "Total": round(item.amount + item_cgst + item_sgst, 2)
            })
    
    df = pd.DataFrame(data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sales', startrow=3)
        
        worksheet = writer.sheets['Sales']
        settings = get_settings()
        
        max_col = len(df.columns)
        
        # Merge header rows
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        
        # Row 1: Shop name
        cell1 = worksheet.cell(row=1, column=1)
        cell1.value = settings.business_name
        cell1.font = Font(size=14, bold=True)
        cell1.alignment = Alignment(horizontal='center')
        
        # Row 2: All Sales
        cell2 = worksheet.cell(row=2, column=1)
        cell2.value = "All Sales"
        cell2.font = Font(size=12, bold=True)
        cell2.alignment = Alignment(horizontal='center')
        
        # Add totals row
        start_row = 5  # data starts from row 5
        data_rows = len(df)
        last_data_row = start_row + data_rows - 1
        total_row = last_data_row + 1
        
        # Column mapping based on dataframe order
        # Invoice No(1), Date(2), Customer(3), GSTIN(4), Product(5), HSN(6), Qty(7), Rate(8), Amount(9), GST%(10), CGST(11), SGST(12), Total(13)
        
        # Add "TOTAL" label in Product Name column (column 5)
        worksheet.cell(row=total_row, column=5).value = "TOTAL"
        worksheet.cell(row=total_row, column=5).font = Font(bold=True)
        
        # Merge TOTAL label across Product Name, HSN, Qty, Rate columns (5-8)
        worksheet.merge_cells(start_row=total_row, start_column=5, end_row=total_row, end_column=8)
        
        # Add formulas for Amount, CGST, SGST, Total
        worksheet.cell(row=total_row, column=9).value = f"=SUM(I{start_row}:I{last_data_row})"
        worksheet.cell(row=total_row, column=9).font = Font(bold=True)
        
        worksheet.cell(row=total_row, column=11).value = f"=SUM(K{start_row}:K{last_data_row})"
        worksheet.cell(row=total_row, column=11).font = Font(bold=True)
        
        worksheet.cell(row=total_row, column=12).value = f"=SUM(L{start_row}:L{last_data_row})"
        worksheet.cell(row=total_row, column=12).font = Font(bold=True)
        
        worksheet.cell(row=total_row, column=13).value = f"=SUM(M{start_row}:M{last_data_row})"
        worksheet.cell(row=total_row, column=13).font = Font(bold=True)
        
        # Auto column width (skipped due to merged cells)
        # for col in worksheet.columns:
        #     max_length = 0
        #     for cell in col:
        #         try:
        #             if cell.value and hasattr(cell, 'column_letter'):
        #                 max_length = max(max_length, len(str(cell.value)))
        #         except:
        #             pass
        #     if max_length > 0:
        #         worksheet.column_dimensions[col[0].column_letter].width = max_length + 2
    
    output.seek(0)
    
    filename = f"all_sales_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
